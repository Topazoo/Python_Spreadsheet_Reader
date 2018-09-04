#!/usr/bin/python

""" Author: Peter Swanson
            pswanson@ucdavis.edu

    Description: A class to make interacting with spreadsheets via 
    the openpyxl library easier

    Version: Python 2.7
    Requirements: openpyxl """

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import os

class Spreadsheet(object):
    """ Class to simplify operating on Excel sheets """

    def __init__(self, filename, load=True):
        """ @filename - The name of the excel file including the extension
            @load - Load the file if true, create it if false """

        # Add extension
        if filename[-5::] != ".xlsx":
            print filename[-5::]
            filename += ".xlsx"

        self.filename = filename

        # If load is False, recreate excel file on every run
        # Else load the file
        if load is True:
            try:
                self.file = load_workbook(filename=filename)
            except IOError:
                print "Failed to load " + filename + "."
                print "Please check if the document exists and try again."
                exit(1)

        else:
            if os.path.isfile(filename):
                os.remove(filename)

            self.file = Workbook()
            self.save()
        
        # Work on default sheet
        self.sheet = self.file.active

    def save(self):
        """ Save the file """

        try:
            self.file.save(filename=self.filename)
        except IOError:
            print "Failed to save " + self.filename + "."
            print "Please check if the document is open and try again."
            exit(1)

    def write(self, content, cell):
        """ Write to a cell
            @content - The content to write
            @cell - The cell to write to """

        self.sheet[cell] = content
        self.save()

    def read_column(self, col, header=True, start_row=1):
        """ Read values from a column
            @col - The column to read
            @header - True if the column has a header. Used as column dict key 
            @start_row - The row to start reading the column from """

        # Column stored in a dict with a single list
        values = []
        column = {}

        row = start_row

        # Header is dict key, values is a list of cell values from the column
        if header:
            column[self.sheet[row][col - 1].value] = values
            row += 1
        else:
            column["None"] = values

        # Store values
        while True:
            cell = self.sheet.cell(row=row, column=col).value

            if cell is None:
                break

            values.append(cell)
            row += 1

        return column

    def read_row(self, row, start_col=1):
        """ Read values from a row
            @row - The row to read
            @start_col - The column to start reading th row from """

        # Row stored in a dict with a single list
        values = []
        obj = {}
        obj['Row: ' + str(row)] = values

        col = start_col

        # Store values
        while True:
            cell = self.sheet.cell(row=row, column=col).value

            if cell is None:
                break

            values.append(cell)
            col += 1

        return obj

    def write_column(self, col, content, start_row=1, bold=False, italics=False):
        """ Write values to a column
            @col - The column to write to
            @content - A list of strings to write 
            @start_row - The row to start writing from
            @bold - True if text should be bolded
            @italics - True if text should be italicized """

        # Column stored in a dict with a single list
        row = start_row

        # Store values
        for value in content:
            try:
                cell = self.sheet.cell(row=row, column=col, value=value.encode('utf-8'))
            except UnicodeEncodeError:
                cell = self.sheet.cell(row=row, column=col, value=value.decode('utf-8'))
            except UnicodeDecodeError:
                cell = self.sheet.cell(row=row, column=col, value="Content Error")

            cell.font = Font(bold=bold, italic=italics)
            row += 1

        self.save()

    def write_row(self, row, content, start_col=1, bold=False, italics=False):
        """ Write values to a row
            @row - The row to write to
            @content - A list of strings to write 
            @start_col - The column to start writing at
            @bold - True if text should be bolded
            @italics - True if text should be italicized """

        # Column stored in a dict with a single list
        col = start_col

        # Store values
        for value in content:
            try:
                cell = self.sheet.cell(row=row, column=col, value=value.encode('utf-8'))
            except UnicodeEncodeError:
                cell = self.sheet.cell(row=row, column=col, value=value.decode('utf-8'))
            except UnicodeDecodeError:
                cell = self.sheet.cell(row=row, column=col, value="Content Error")

            cell.font = Font(bold=bold, italic=italics)
            col += 1

        self.save()

    def append_row(self, row, content, bold=False, italics=False):
        """ Append content to row
            @row - The row to append to
            @content - A list of strings to write 
            @bold - True if text should be bolded
            @italics - True if text should be italicized """

        col = 1

        # Search for next empty column in row
        while True:
            cell = self.sheet.cell(row=row, column=col).value

            if cell is None:
                break

            col += 1

        # Write after
        self.write_row(row, content, col, bold, italics)

    def append_column(self, col, content, bold=False, italics=False):
        """ Append content to column 
            @col - The column to append to
            @content - A list of strings to write 
            @bold - True if text should be bolded
            @italics - True if text should be italicized """

        row = 1

        # Search for next empty row in column
        while True:
            cell = self.sheet.cell(row=row, column=col).value

            if cell is None:
                break

            row += 1

        # Write after
        self.write_column(col, content, row, bold, italics)

    def create_sheets(self, values):
        """ Create sheets to write to
            @values - A list of strings to title sheets """

        for value in values:
            self.file.create_sheet(value)

        self.save()